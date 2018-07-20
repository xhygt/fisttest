# encoding =utf-8
# openpyxl 只支持 .xlsx 格式

import openpyxl
from openpyxl.styles import  Border,Side,Font
import time

class ParseExcel(object):

    def __init__(self):
        self.workbook = None
        self.excelFile = None
        self.font = Font(color = None) # 设置字体的颜色
        # 颜色对应的 RGB 值
        self.RGBDict = {'red': 'FFFF3030', 'green':'FF008B00'}

    def loadWorkBook(self,excelPathAndName):
        # 将 Excel 文件加载到内存，并获取其 workbook 对象
        try:
            self.workbook = openpyxl.load_workbook(excelPathAndName)
        except Exception as e:
            raise e
        self.excelFile = excelPathAndName
        return  self.workbook

    def getSheetByname(self,sheetName):
        # 根据 sheet 名获取该 sheet 对象
        try:
            sheet = self.workbook.get_sheet_by_name(sheetName)
            return  sheet
        except Exception as e:
            raise e

    def getSheetByIndex(self,sheetIndex):
        # 根据 sheet 的索引号获取该 sheet 对象
        try:
            sheetname = self.workbook.get_sheet_names()
        except Exception as e:
            raise e
        sheet = self.workbook.get_sheet_by_name(sheetname[sheetIndex])
        return sheet

    def getRowsNumber(self,sheet):
        # 获取 sheet 中有数据区域的结束行号
        return  sheet.max_row

    def getColsNumber(self,sheet):
        # 获取 sheet 中有数据区域的结束列号
        return  sheet.max_column

    def getStartRowNumber(self,sheet):
        # 获取 sheet 中有数据区域的开始行号
        return  sheet.min_row

    def getStartColNumber(self,sheet):
        # 获取 sheet 中有数据区域的开始列好：
        return  sheet.min_column

    def getRow(self,sheet,rowNo):

        # 获取 sheet 中某一行，返回的是这一行所有的数据内容组成的 tuple,
        # 下表从 1 开始，sheet.rows[1]表示第一行
        try:
            #print(sheet.rows[rowNo])
             for cell in list(sheet.rows)[rowNo-1]:
                print(cell.value)
        except Exception as e:
            raise e

    def getRows(self,sheet):
        # 遍历表格中的数据
        try:
            for row in sheet.iter_rows():
                for cell in row:
                    print(cell.value)
        except Exception as e:
            raise e

    def getCellOfValue(self,sheet,coordinate = None,
                       rowNo = None,colsNo = None):
        # 根据单元格所在的位置索引获取该单元格中的值，下标从1开始
        # sheet.cell(row =1 ,column =1 ).value,
        # 表示  excel 中第一行第一列的值
        if coordinate !=None:
            try:
                return  sheet.cell(coordinate = coordinate).value
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and \
            colsNo is not None:
            try:
                return sheet.cell(row = rowNo,cols = colsNo).value
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell !")

    def getCellOfObject(self,sheet,coordinate = None,
                        rowNo = None,colsNo =None):
        # 获取某个单元格的对象，可以根据单元格所在位置的数字索引，
        # 也可以直接根据Excel 中单元格的编码及坐标
        # 如 getCellObject(sheet,coordinate = 'A1') or
        # getCellObject(sheet,rowNo = 1,colsNo = 2)
        if coordinate != None:

            try:
                return  sheet.cell(coordinate = coordinate)
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is None and \
            colsNo is not None:
            try:
                return sheet.cell(row = rowNo,column = colsNo)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell !")

    def writeCell(self,sheet,content,coordinate = None,
                  rowNo = None,colsNo = None,style = None):
        # 根据单元格在 Excel 中的编码坐标或者数字索引坐标向单元格中写入数据，
        # 下标从 1 开始，参数 style 表示字体的颜色的名字，比如 red,green
        if coordinate is not None:
            try:
                sheet.cell(coordinate = coordinate).value = content
                if style is not None:
                    sheet.cell(coordinate = coordinate).\
                        font = Font(color = self.RGBDict[style])
                    self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and \
            colsNo is not None:
            try:
                sheet.cell(row = rowNo,column = colsNo).value = content
                if style:
                    sheet.cell(row = rowNo,column = colsNo).\
                        font = Font(color = self.RGBDict[style])
                    self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell !")

    def writeCellCurrentTime(self,sheet,coordinate = None,
                             rowNo = None,colsNo =None):
        # 写入当前的时间，下标从 1开始
        now = int(time.time())
        timeArray = time.localtime(now)
        currentTime = time.strftime("%Y-%m-%d %H:%M:%S",timeArray)
        if coordinate is not None:
            try:
                sheet.cell(coordinate = coordinate).value = currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None \
            and colsNo is not None:
            try:
                sheet.cell(row = rowNo,column = colsNo).value = currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell !")


if __name__ == '__main__':
    # 测试代码

    pe = ParseExcel()
    pe.loadWorkBook(u'/Users/dingyq/pyauto/testdata/test.xlsx')
    print("通过名称获取 sheet 对象的名字:", \
          pe.getSheetByname(u'联系人'))
    print("通过 index 序号获取 sheet 对象的名字：", \
          pe.getSheetByIndex(0).title)
    sheet = pe.getSheetByIndex(0)
    #print(pe.getColsNumber(sheet))
    #print(pe.getRowsNumber(sheet))
    print(pe.getRow(sheet,1))

    pe.writeCell(sheet,u'我爱祖国', rowNo = pe.getRowsNumber(sheet)+1,
                 colsNo = pe.getStartColNumber(sheet))
    pe.writeCellCurrentTime(sheet,rowNo = pe.getRowsNumber(sheet),
                 colsNo = pe.getStartColNumber(sheet)+1)
    #print(pe.getRows(sheet))

